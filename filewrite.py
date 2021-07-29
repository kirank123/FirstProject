# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}' + " how are you ?")  # Press Ctrl+F8 to toggle the breakpoint.
    print("I'm fine " + name + ". Thank You!")


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('Kiran Krishnan')

    import openpyxl

    path = r"C:\Users\kiran\OneDrive\Desktop\newdata.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    wb = openpyxl.Workbook()
    sht = wb.active
    sht.title = "Kiran"
    wb.create_sheet(index=-1, title="Krishnan")
    wb.create_sheet(index=0, title="Python")

    sht.sheet_properties.tabColor = "1072BA"

    for i in range(sheet.max_row):
        for j in range(sheet.max_column):
            sht.cell(row=i+1, column=j+1).value = sheet.cell(row=i+1, column=j+1).value


    wb.save(r"C:\Users\kiran\OneDrive\Desktop\newdata1.xlsx")
