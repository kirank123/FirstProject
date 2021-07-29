# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name} how are you ?')  # Press Ctrl+F8 to toggle the breakpoint.
    print("I'm fine " + name + ". Thank You!")
    print(f"I'm fine {name}. Thank You!")


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('Kiran Krishnan')

    import openpyxl

#    path = r"C:\Users\kiran\OneDrive\Desktop\new.xlsx"
#    workbook = openpyxl.load_workbook(path)
#    sheet = workbook.active
#    cell = sheet.cell(row=2, column=1)

#    for i in range(sheet.max_row):
#        for j in range(sheet.max_column):
#            print(sheet.cell(row=i+1, column=j+1).value)


wb = openpyxl.Workbook()
sht = wb.active
sht.title = "Kiran"
wb.create_sheet(index=1, title="Krishnan")
wb.create_sheet(index=2, title="Python")
sht['A1'].value = "Kiran Krishnan"
sht['A2'].value = "Mallisseril(H)"
sht['A3'].value = "Karimannoor P.O"
sht['A4'].value = "Thodupuzha"
sht['A5'].value = "Idukki"
sht['A6'].value = "Kerala"

wb.save(r"C:\Users\kiran\OneDrive\Desktop\newdata.xlsx")
